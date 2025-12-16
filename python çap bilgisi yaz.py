import openpyxl
from openpyxl.utils import get_column_letter
import re

# Excel dosyasını aç
# wb = openpyxl.load_workbook('ExcelReport (6).xlsx')
wb = openpyxl.load_workbook('C:\\Users\\yak\\Desktop\\Python\\ExcelReport (6).xlsx')
ws = wb['rphaldemtuketim']

# Yeni bir sütun ekleme (örneğin 'ÇAP' sütunu)
new_column_index = ws.max_column + 1
ws.cell(row=1, column=new_column_index, value='ÇAP')

# Verileri oku ve çap bilgilerini yeni sütuna yaz
for row in range(2, ws.max_row + 1):  # min_row=2 ilk satırı başlık olarak atlıyoruz
    product_definition = ws.cell(row=row, column=7).value  # GİRİŞ ÜRÜN TANIMI sütunu
    
    if product_definition:
        # Çap bilgisini regular expression ile çıkart
        match = re.search(r'(\d+\.\d+)MM', product_definition)
        if match:
            cap = float(match.group(1))
            ws.cell(row=row, column=new_column_index, value=cap)
        else:
            ws.cell(row=row, column=new_column_index, value=None)  # Çap bilgisi yoksa boş bırak
    else:
        ws.cell(row=row, column=new_column_index, value=None)  # Ürün tanımı yoksa boş bırak

# Değişiklikleri kaydet
wb.save('ornek_dosya_cizgi_bilgi_eklenmis.xlsx')
print("Çap bilgileri Excel dosyasına yazıldı.")