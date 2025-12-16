import openpyxl

# Yeni bir Excel dosyası oluştur
wb = openpyxl.Workbook()

# Aktif çalışma sayfasını al
ws = wb.active

# Çalışma sayfasına bazı verileri yaz
ws['A1'] = 'İsim'
ws['B1'] = 'Yaş'
ws['C1'] = 'Şehir'

# Daha fazla veri ekle
data = [
    ['Ali', 25, 'Ankara'],
    ['Ayşe', 30, 'İstanbul'],
    ['Mehmet', 22, 'İzmir']
]

for row in data:
    ws.append(row)

# Excel dosyasını kaydet
wb.save('ornek_dosya.xlsx')
print("Veriler Excel dosyasına yazıldı.")

# Şimdi aynı dosyadan verileri okuyalım
wb = openpyxl.load_workbook('ornek_dosya.xlsx')
ws = wb.active

# Çalışma sayfasındaki tüm hücreleri oku
for row in ws.iter_rows(values_only=True):
    print(row)